"""Vistas para la bitácora de auditoría general (Logs).

Control de acceso:
  - AuditoriaLogsView               → AUDITOR, ADMINISTRADOR, DIRECTOR
  - ExportarAuditoriaView           → AUDITOR, ADMINISTRADOR, DIRECTOR
"""
import csv
import io
import os
from datetime import date, timedelta
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db import connection
from django.http import HttpResponse
from django.conf import settings
from api.permissions import IsAuditorOrAdmin
from api.utils.auditoria import registrar_evento, log_descarga


def _build_logs_query(request):
    """Construye la consulta SQL de logs con filtros y retorna (sql, params)."""
    hoy = date.today()
    fecha_desde = request.query_params.get('desde', str(hoy - timedelta(days=30)))
    fecha_hasta = request.query_params.get('hasta', str(hoy))
    accion = request.query_params.get('accion', '').strip().upper()
    busqueda = request.query_params.get('q', '').strip()

    where_clauses = [
        "DATE((al.fecha_hora AT TIME ZONE 'UTC') AT TIME ZONE 'America/Caracas') BETWEEN %s AND %s"
    ]
    params = [fecha_desde, fecha_hasta]

    if accion:
        where_clauses.append("al.accion = %s")
        params.append(accion)

    if busqueda:
        where_clauses.append("(COALESCE(au.username,'') ILIKE %s OR al.descripcion ILIKE %s OR al.accion ILIKE %s)")
        params.extend([f'%{busqueda}%', f'%{busqueda}%', f'%{busqueda}%'])

    where_sql = ' AND '.join(where_clauses)

    sql = f"""
        SELECT
            al.id_log,
            al.accion,
            al.descripcion,
            ((al.fecha_hora AT TIME ZONE 'UTC') AT TIME ZONE 'America/Caracas') AS fecha_hora,
            al.ip_address,
            COALESCE(au.username, 'Sistema') AS usuario,
            COALESCE(al.sistema, 'SISTEMA') AS departamento
        FROM auditoria_logs al
        LEFT JOIN auth_user au ON al.id_usuario = au.id
        WHERE {where_sql}
        ORDER BY al.fecha_hora DESC
    """
    return sql, params, fecha_desde, fecha_hasta


class AuditoriaLogsView(APIView):
    """Lista de eventos del sistema para Auditoría con filtros por fecha y acción."""
    permission_classes = [IsAuditorOrAdmin]

    def get(self, request):
        sql, params, _, _ = _build_logs_query(request)
        limit = int(request.query_params.get('limit', 200))
        sql += f" LIMIT {limit}"

        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Serializar datetimes
        for row in data:
            if row.get('fecha_hora'):
                row['fecha_hora'] = row['fecha_hora'].isoformat()

        return Response(data)


def _build_pdf_letterhead(elements, styles, titulo, desde, hasta):
    """Agrega el membrete institucional al principio de los elementos PDF."""
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    import base64
    import io
    from api.utils import get_logo_base64
    from reportlab.platypus import Image

    # Formatear fechas
    desde_fmt = desde[8:10] + '-' + desde[5:7] + '-' + desde[0:4] if len(desde) >= 10 else desde
    hasta_fmt = hasta[8:10] + '-' + hasta[5:7] + '-' + hasta[0:4] if len(hasta) >= 10 else hasta

    logo_left = None
    logo_right = None

    dem_left_base64 = get_logo_base64('dem-2.png')
    dem_right_base64 = get_logo_base64('dem.png')

    if dem_left_base64 and dem_right_base64:
        try:
            logo_left = Image(io.BytesIO(base64.b64decode(dem_left_base64)), width=3.5*cm, height=1.2*cm)
            logo_right = Image(io.BytesIO(base64.b64decode(dem_right_base64)), width=1.4*cm, height=1.4*cm)
        except Exception:
            pass

    # Fallback al sistema de archivos local si falla Base64
    if not logo_left or not logo_right:
        logo_left_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'dem-2.png')
        if not os.path.exists(logo_left_path):
            logo_left_path = os.path.join(settings.BASE_DIR, '..', 'frontend', 'src', 'assets', 'img', 'dem-2.png')
            
        logo_right_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'dem.png')
        if not os.path.exists(logo_right_path):
            logo_right_path = os.path.join(settings.BASE_DIR, '..', 'frontend', 'src', 'assets', 'img', 'dem.png')

        if os.path.exists(logo_left_path) and os.path.exists(logo_right_path):
            try:
                logo_left = Image(logo_left_path, width=3.5*cm, height=1.2*cm)
                logo_right = Image(logo_right_path, width=1.4*cm, height=1.4*cm)
            except Exception:
                pass

    if logo_left and logo_right:
        header_data = [[
            logo_left, 
            Paragraph(
                '<div align="center"><b>DIRECCIÓN EJECUTIVA DE LA MAGISTRATURA</b><br/>'
                '<font size="8">SISTEMA INSTITUCIONAL GENERAL - DEM</font><br/>'
                f'<font size="7" color="grey">Período: {desde_fmt} al {hasta_fmt}</font></div>',
                styles['Normal']
            ),
            logo_right
        ]]
        header_table = Table(header_data, colWidths=['20%', '60%', '20%'])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#1e3a5f')),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ]))
    else:
        header_data = [[
            Paragraph(
                '<div align="center"><b>DIRECCIÓN EJECUTIVA DE LA MAGISTRATURA</b><br/>'
                f'<font size="8">SISTEMA GENERAL | Período: {desde_fmt} al {hasta_fmt}</font></div>',
                styles['Normal']
            )
        ]]
        header_table = Table(header_data, colWidths=['100%'])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#1e3a5f')),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ]))
    elements.append(header_table)
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f'<b>{titulo}</b>', styles['Title']))
    elements.append(Spacer(1, 8))


class ExportarAuditoriaView(APIView):
    """Exporta la bitácora de auditoría en CSV, Excel o PDF con membrete institucional."""
    permission_classes = [IsAuditorOrAdmin]

    def get(self, request):
        formato = request.query_params.get('formato', 'csv').lower()
        sql, params, fecha_desde, fecha_hasta = _build_logs_query(request)

        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        fecha_desde_fmt = fecha_desde[8:10] + '-' + fecha_desde[5:7] + '-' + fecha_desde[0:4] if len(fecha_desde) >= 10 else fecha_desde
        fecha_hasta_fmt = fecha_hasta[8:10] + '-' + fecha_hasta[5:7] + '-' + fecha_hasta[0:4] if len(fecha_hasta) >= 10 else fecha_hasta

        columns = ['ID', 'Departamento', 'Acción', 'Descripción', 'Fecha/Hora', 'IP', 'Usuario']
        titulo = f"Bitácora de Eventos — {fecha_desde_fmt} al {fecha_hasta_fmt}"
        filename = f"auditoria_{fecha_desde_fmt}_{fecha_hasta_fmt}"

        formatted_rows = []
        for row in rows:
            row = list(row)
            if row[3]:
                try:
                    row[3] = row[3].strftime('%d-%m-%Y %H:%M:%S')
                except Exception:
                    row[3] = str(row[3])
            reordered = [row[0], row[6], row[1], row[2], row[3], row[4], row[5]]
            formatted_rows.append(reordered)

        if formato == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            response.write('\ufeff')
            writer = csv.writer(response)
            writer.writerow(columns)
            for row in formatted_rows:
                writer.writerow([str(c) if c is not None else '' for c in row])
            log_descarga(request, "AUDITORIA_CSV", f"{filename}.csv")
            return response

        elif formato == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = Workbook()
                ws = wb.active
                ws.title = "Bitácora"

                header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                for row_idx, row in enumerate(formatted_rows, start=2):
                    for col_idx, value in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')

                for column in ws.columns:
                    max_len = max((len(str(cell.value or '')) for cell in column), default=10)
                    ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 60)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
                log_descarga(request, "AUDITORIA_EXCEL", f"{filename}.xlsx")
                return response
            except ImportError:
                return Response({'detail': 'openpyxl no está instalado.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        elif formato == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib.enums import TA_CENTER

                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                         leftMargin=10, rightMargin=10, topMargin=20, bottomMargin=20)
                styles = getSampleStyleSheet()
                styles['Title'].alignment = TA_CENTER
                elements = []

                _build_pdf_letterhead(elements, styles, titulo, fecha_desde, fecha_hasta)

                pdf_columns = ['Dpto.', 'Usuario', 'Acción', 'Descripción', 'Fecha/Hora', 'IP']
                pdf_rows = []
                for r in formatted_rows:
                    row_data = []
                    for c in [r[1], r[6], r[2], str(r[3]), r[4], r[5]]:
                        val = str(c) if c is not None else ''
                        row_data.append(Paragraph(f'<font size="6">{val}</font>', styles['Normal']))
                    pdf_rows.append(row_data)

                table_data = [pdf_columns] + pdf_rows
                col_widths = [60, 90, 80, 362, 80, 80]

                t = Table(table_data, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('WORDWRAP', (2, 1), (2, -1), 'LTR'),
                ]))
                elements.append(t)

                elements.append(Spacer(1, 12))
                elements.append(Paragraph(
                    f'<font size="7" color="grey">Documento generado automáticamente. '
                    f'Total de eventos: {len(formatted_rows)}. '
                    f'Dirección Ejecutiva de la Magistratura.</font>',
                    styles['Normal']
                ))

                doc.build(elements)
                buffer.seek(0)
                response = HttpResponse(buffer.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
                log_descarga(request, "AUDITORIA_PDF", f"{filename}.pdf")
                return response
            except ImportError:
                return Response({'detail': 'reportlab no está instalado.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'detail': 'Formato no soportado. Use: csv, excel, pdf'}, status=status.HTTP_400_BAD_REQUEST)
